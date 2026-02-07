from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def create_excel(data: dict, file_id: str) -> str:
    """Create Excel file in YuktaOne Deal/Placement format - exact template"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Deal & Placements"
    
    # Styles
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    header_font = Font(bold=True, size=10)
    label_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    deal = data.get("deal", {})
    placements = data.get("placements", [])
    
    row = 1
    
    # === DEAL CREATION FORM FIELDS HEADER ===
    headers = ["Deal Creation Form Fields", "Deal Letter Agreement Fields", "Comments"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1
    
    # Deal fields - EXACTLY 22 fields as per template
    deal_rows = [
        ("Deal Name", "N/A", "Need to fill by User."),
        ("Deal Type", deal.get("deal_type", ""), "Should take Agency or Direct based on Agency field"),
        ("Advertiser Name", deal.get("advertiser_name", ""), ""),
        ("Client Name", "N/A", "Will auto-reflect, based on Advertiser Name"),
        ("Brand Name", deal.get("brand_name", ""), "If empty in Sheet, then user needs to select"),
        ("Product Category", deal.get("product_category", ""), "If empty in Sheet, then user needs to select"),
        ("Agency Name", deal.get("agency_name", ""), ""),
        ("Tournament Name", deal.get("tournament_name", ""), ""),
        ("Sales Person", deal.get("sales_person", ""), "From Star India Contact Person"),
        ("Plant", "N/A", "Will be auto populated based on Sales Person"),
        ("Zone", "N/A", "Will be auto populated based on Sales Person"),
        ("Sales Group", "N/A", "Need to fill by User."),
        ("Start Date", deal.get("start_date", ""), ""),
        ("End Date", deal.get("end_date", ""), ""),
        ("Deal Currency", "N/A", "Need to fill by User."),
        ("Execution Currency", deal.get("execution_currency", "INR"), "By default INR in YuktaOne"),
        ("Currency Conversion Rate", "N/A", "Need to fill by User if Deal Currency is other than INR."),
        ("Booked Revenue", deal.get("booked_revenue", ""), ""),
        ("Discount Type", "N/A", "Disable"),
        ("Discount", "N/A", "Disable"),
        ("Booked Revenue in Execution Currency After Discount", "N/A", "Auto-populated field."),
        ("Booked Revenue in Deal Currency After Discount", "N/A", "Auto-populated field."),
    ]
    
    for field, value, comment in deal_rows:
        ws.cell(row=row, column=1, value=field).border = thin_border
        ws.cell(row=row, column=2, value=value if value else "").border = thin_border
        ws.cell(row=row, column=3, value=comment).border = thin_border
        row += 1
    
    # Empty rows
    row += 2
    
    # === PLACEMENT CREATION FORM FIELDS ===
    for idx, placement in enumerate(placements):
        # Header for each placement
        headers = ["Placement Creation Form Fields", "Deal Letter Agreement Fields", "Comments"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1
        
        # Placement fields - EXACTLY 30 fields as per template
        placement_rows = [
            ("Billing Type", placement.get("billing_type", "Standard"), "Standard by default"),
            ("Ad Server Name", "N/A", "Based on Ad Format/Ad Assets"),
            ("Start Date", placement.get("start_date") or deal.get("start_date", ""), ""),
            ("End Date", placement.get("end_date") or deal.get("end_date", ""), ""),
            ("Tournament Name", placement.get("tournament_name") or deal.get("tournament_name", ""), "Auto-populated based on Deal"),
            ("Buy Type", placement.get("buy_type", ""), "Need to select by User."),
            ("Ad Format", placement.get("ad_format", ""), ""),
            ("Ad Duration", placement.get("ad_duration", ""), ""),
            ("Content Type", placement.get("content_type", ""), ""),
            ("Spot Type", placement.get("spot_type", "Original"), "Original by default"),
            ("Platform", placement.get("platform", ""), ""),
            ("Match", placement.get("match", ""), "Need to select by User."),
            ("Creative ID", "N/A", ""),
            ("RO Number", "N/A", "Need to select by User."),
            ("Stream", "N/A", "Need to select by User."),
            ("Material Number", "N/A", "Need to select by User."),
            ("Booked Quantity", placement.get("booked_quantity", ""), ""),
            ("Bonus Quantity", "N/A", ""),
            ("Total Quantity", "N/A", "Auto-populated field."),
            ("Rate", placement.get("rate", ""), ""),
            ("Booked Revenue", "N/A", "Auto-populated field."),
            ("Campaign Name", "N/A", ""),
            ("Targeting", placement.get("targeting", ""), ""),
            ("Campaign Manager", "N/A", ""),
            ("Client", deal.get("advertiser_name", ""), "Auto-populated based on Deal"),
            ("Product Category", deal.get("product_category", ""), "Auto-populated based on Deal"),
            ("Brand Name", placement.get("brand_name") or deal.get("brand_name", ""), ""),
            ("Targeting Comments", "N/A", ""),
            ("Placement Comments", "N/A", ""),
            ("Placement Name", "N/A", "Auto-populated field."),
        ]
        
        for field, value, comment in placement_rows:
            ws.cell(row=row, column=1, value=field).border = thin_border
            ws.cell(row=row, column=2, value=value if value else "").border = thin_border
            ws.cell(row=row, column=3, value=comment).border = thin_border
            row += 1
        
        # Add separator if more placements
        if idx < len(placements) - 1:
            row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 55
    
    # Save
    output_path = f"output_{file_id}.xlsx"
    wb.save(output_path)
    
    return output_path
